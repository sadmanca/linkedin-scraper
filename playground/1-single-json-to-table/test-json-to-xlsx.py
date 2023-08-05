import json
import openpyxl
import os

# Read the headers from a text file
with open('table-headers.txt', 'r') as file:
    headers = [line.strip() for line in file]

# Create a new workbook and select the active worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Write the headers to the worksheet
for i, header in enumerate(headers):
    worksheet.cell(row=1, column=i+1, value=header)

# Loop through each JSON file in the directory
for row_num, filename in enumerate(os.listdir('.'), start=2):
    if filename.endswith('.json'):
        # Load the JSON data from the file
        with open(filename, 'r') as jsonfile:
            data = json.load(jsonfile)

        # Extract the relevant data from the JSON object
        row = []
        for header in headers:
            if header == 'titles':
                # Extract the value of the 'title' key from each item in the 'experience' list
                titles = [
                    f"{item.get('title', '')} at {item.get('companyName', '')}" +
                    (f" ({item.get('locationName', '')})" if item.get('locationName', '') else '')
                    for item in data.get('experience', [])
                ]
                if titles:
                    row.append(', '.join(titles))
                else:
                    row.append('')
                    
            elif header == 'job descriptions':
                descriptions = [f"{item.get('description', 'NO_DESCRIPTION')} at {item.get('companyName', '')}" for item in data.get('experience', [])]
                if descriptions:
                    row.append(', '.join(descriptions))
                else:
                    row.append('')    
                                
            elif header == 'name':
                # Concatenate the values of the 'firstName' and 'lastName' keys with a space separator
                row.append(data.get('firstName', '') + ' ' + data.get('lastName', ''))
                
            elif header == 'public_id':
                # Append 'https://www.linkedin.com/in/' followed by the value of the 'public_id' key to the 'row' list
                row.append('https://www.linkedin.com/in/' + data.get(header, ''))
                
            else:
                # Use the value of the header as-is
                row.append(data.get(header, ''))

        # Write the data to the worksheet
        for col_num, value in enumerate(row, start=1):
            if isinstance(value, list) and not value:
                # Skip writing to the cell if the value is an empty list
                continue
            worksheet.cell(row=row_num, column=col_num, value=value)

# Save the workbook to an XLSX file
workbook.save('test-output.xlsx')