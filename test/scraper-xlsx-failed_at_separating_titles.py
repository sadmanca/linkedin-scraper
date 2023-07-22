import json
import openpyxl
import os

# Read the headers from a text file
with open('headers.txt', 'r') as file:
    headers = [line.strip() for line in file]

# Create a new workbook and select the active worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Write the headers to the worksheet
for i, header in enumerate(headers):
    worksheet.cell(row=1, column=i+1, value=header)

# Loop through each JSON file in the directory
for row_num, filename in enumerate(os.listdir('.'), start=2):
    if filename.endswith('.json'):
        # Load the JSON data from the file
        with open(filename, 'r') as jsonfile:
            data = json.load(jsonfile)

        # Extract the relevant data from the JSON object
        row = []
        for header in headers:
            if header == 'title':
                # Extract the value of the 'title' and 'companyName' keys from each item in the 'experience' list
                titles = [f"{item.get('title', '')} at {item.get('companyName', '')}" for item in data.get('experience', [])]
                
                # Use a flag variable to keep track of the first instance of 'title' with 'companyName'
                current_position_written = False
                
                if titles:
                    for title in titles:
                        if not current_position_written:
                            # Write the first instance of 'title' with 'companyName' under the header 'currentPosition'
                            row.append(title)
                            current_position_written = True
                        else:
                            # Write all other instances of 'title' with 'companyName' under the header 'previousExperience'
                            row.append(title)
                else:
                    row.append('')
            else:
                # Use the value of the header as-is
                row.append(data.get(header, ''))

        # Write the data to the worksheet
        for col_num, value in enumerate(row, start=1):
            if isinstance(value, list) and not value:
                # Skip writing to the cell if the value is an empty list
                continue
            worksheet.cell(row=row_num, column=col_num, value=value)

# Save the workbook to an XLSX file
workbook.save('output.xlsx')